"""Read catalogue spreadsheets into validated series.

One sheet is one series. The sheet name carries the identity of the product in the form
``<product> | <article no.>``; cells ``A1`` and ``B1`` carry the axis labels in the form
``name [unit]``.

Column A holds the y axis and column B the x axis. The inversion is deliberate and worth
stating once: the polynomial this tool eventually fits answers "which valve setting produces
this kv", so kv is the free variable and the setting is what gets computed.

Failure granularity follows the contract. A sheet that cannot describe a series is reported
against that sheet and the rest of the workbook continues, because one malformed sheet in a
catalogue of fourteen is not a reason to import nothing. A problem with the file itself —
absent, unreadable, or missing the sheet the caller asked for — stops the operation.

Nothing here reads the clock: ``imported_at`` arrives as an argument, which is what keeps a
stored catalogue byte-identical between runs.
"""

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from hydrofit.errors import HydrofitError
from hydrofit.models import AxisSpec, DataKind, Series, SourceRef

# A label is "name [unit]". The unit pattern forbids brackets of its own so that a malformed
# label cannot be salvaged into something plausible-looking by a greedy match.
_LABEL = re.compile(r"^\s*(?P<name>\S.*?)\s*\[\s*(?P<unit>[^\[\]]+?)\s*\]\s*$")

_SEPARATOR = "|"

# The catalogue puts the points in the first two columns. Pinning the width keeps a stray note
# parked in column C from turning a healthy sheet into a reported problem.
_POINT_COLUMNS = 2


@dataclass(frozen=True, slots=True)
class SheetProblem:
    """One sheet that could not describe a series, and the reason.

    Attributes:
        sheet: Sheet name as it appears in the workbook.
        message: Why the sheet was rejected, phrased for the person who has to fix it.
    """

    sheet: str
    message: str


@dataclass(frozen=True, slots=True)
class ImportResult:
    """Everything one workbook yielded: the series that survived, and the sheets that did not.

    Both halves matter. Reporting only the series would hide a sheet that silently stopped
    being imported; reporting only a count would not say which one.

    Attributes:
        series: The series built from the workbook, in sheet order.
        problems: The sheets that could not become series, in sheet order.
    """

    series: tuple[Series, ...]
    problems: tuple[SheetProblem, ...]


def split_sheet_name(name: str) -> tuple[str, str]:
    """Split a sheet name into the product and the article number.

    Args:
        name: Sheet name, expected in the form ``<product> | <article no.>``.

    Returns:
        The product and the article number, both stripped.

    Raises:
        HydrofitError: If the separator is missing, appears more than once, or either side of
            it is blank.
    """
    parts = name.split(_SEPARATOR)
    if len(parts) != 2:
        raise HydrofitError(
            f"sheet name {name!r} is not the form <product> {_SEPARATOR} <article no.>"
        )
    product, article_no = (part.strip() for part in parts)
    if not product:
        raise HydrofitError(f"sheet name {name!r} has no product before the separator")
    if not article_no:
        raise HydrofitError(
            f"sheet name {name!r} has no article number after the separator"
        )
    return product, article_no


def parse_label(value: object, cell: str) -> AxisSpec:
    """Read an axis label of the form ``name [unit]``.

    Args:
        value: Cell value as the spreadsheet returned it.
        cell: Cell reference, used to say where the problem is.

    Returns:
        The axis the label describes. The brackets are not stored — they belong to the display
        layer, and `AxisSpec.label` puts them back.

    Raises:
        HydrofitError: If the cell is not text, or the text carries no unit in brackets.
    """
    if not isinstance(value, str):
        raise HydrofitError(f"cell {cell} does not hold an axis label: {value!r}")
    match = _LABEL.match(value)
    if match is None:
        raise HydrofitError(
            f"cell {cell} is not the form name [unit]: {value!r} — a dimensionless "
            f"quantity carries a dash, never an empty unit"
        )
    return AxisSpec(match.group("name"), match.group("unit"))


def parse_number(value: object, cell: str) -> float:
    """Read one numeric cell, tolerating the decimal comma.

    Args:
        value: Cell value as the spreadsheet returned it.
        cell: Cell reference, used to say where the problem is.

    Returns:
        The value as a float.

    Raises:
        HydrofitError: If the cell holds anything that is not a number.
    """
    # bool is checked before the numeric case because it subclasses int: TRUE in a data cell
    # would otherwise slip through as 1.0, and the series would look healthy while carrying a
    # value nobody measured.
    if isinstance(value, bool):
        raise HydrofitError(f"cell {cell} holds a true/false value, not a number")
    if isinstance(value, int | float):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", ".")
        try:
            return float(text)
        except ValueError as exc:
            raise HydrofitError(f"cell {cell} is not a number: {value!r}") from exc
    raise HydrofitError(f"cell {cell} is not a number: {value!r}")


def _read_points(rows: list[tuple[object, ...]]) -> tuple[list[float], list[float]]:
    """Read the data rows of a sheet into y and x values.

    Args:
        rows: Cell values from row 2 onwards, two columns wide.

    Returns:
        The y values (column A) and the x values (column B), paired by position.

    Raises:
        HydrofitError: If the sheet holds no points, a row is half empty, or a cell is not a
            number.
    """
    y_values: list[float] = []
    x_values: list[float] = []
    for offset, row in enumerate(rows):
        number = offset + 2  # row 1 holds the labels
        y_cell, x_cell = row[0], row[1]
        # A wholly empty row is spreadsheet slack — trailing padding, a visual gap — and is
        # skipped. A half-empty one is not: it means a point lost one of its coordinates.
        if y_cell is None and x_cell is None:
            continue
        if y_cell is None or x_cell is None:
            raise HydrofitError(f"row {number} has a value in only one column")
        y_values.append(parse_number(y_cell, f"A{number}"))
        x_values.append(parse_number(x_cell, f"B{number}"))
    if not y_values:
        raise HydrofitError("the sheet holds no points")
    return y_values, x_values


def _read_sheet(
    *,
    rows: list[tuple[object, ...]],
    name: str,
    file: str,
    imported_at: str,
    kind: DataKind,
) -> Series:
    """Turn the cells of one sheet into a series.

    Args:
        rows: Every row of the sheet, two columns wide.
        name: Sheet name, which carries the product and the article number.
        file: File name to record in the source reference.
        imported_at: ISO-8601 timestamp to record in the source reference.
        kind: Whether these points are raw readings or generated from a curve.

    Returns:
        The series the sheet describes.

    Raises:
        HydrofitError: If the name, the labels, or the points do not describe a series. The
            caller turns this into a problem reported against this sheet.
    """
    product, article_no = split_sheet_name(name)
    if not rows:
        raise HydrofitError("the sheet is empty")

    header = rows[0]
    y_axis = parse_label(header[0], "A1")
    x_axis = parse_label(header[1], "B1")
    y_values, x_values = _read_points(rows[1:])

    return Series(
        product=product,
        article_no=article_no,
        x_axis=x_axis,
        y_axis=y_axis,
        x=tuple(x_values),
        y=tuple(y_values),
        kind=kind,
        source=SourceRef(file=file, sheet=name, imported_at=imported_at),
    )


def import_workbook(
    path: Path | str,
    *,
    imported_at: str,
    kind: DataKind,
    sheet: str | None = None,
) -> ImportResult:
    """Read a catalogue workbook into series.

    Args:
        path: Path to the ``.xlsx`` file.
        imported_at: ISO-8601 timestamp recorded on every series read here. The importer never
            reads the clock; the caller decides what time it is.
        kind: Whether these points are raw readings or generated from a curve. Passed in by the
            caller for now; classification by shape arrives with the heuristic.
        sheet: Name of a single sheet to import. ``None`` imports every sheet.

    Returns:
        The series that could be built and the sheets that could not.

    Raises:
        HydrofitError: If the file is absent or unreadable, holds no sheets, or does not hold
            the sheet that was asked for.
    """
    path = Path(path)
    if not path.is_file():
        raise HydrofitError(f"no such spreadsheet: {path}")

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        # openpyxl reports a wrong or damaged file through several unrelated exception types,
        # so the class is not a reliable filter here; what the user needs is the message.
        raise HydrofitError(f"cannot read {path.name} as a spreadsheet: {exc}") from exc

    try:
        names: list[str] = list(workbook.sheetnames)
        if not names:
            raise HydrofitError(f"{path.name} holds no sheets")
        if sheet is not None:
            if sheet not in names:
                raise HydrofitError(f"{path.name} has no sheet named {sheet!r}")
            names = [sheet]

        series: list[Series] = []
        problems: list[SheetProblem] = []
        for name in names:
            rows = [
                tuple(row)
                for row in workbook[name].iter_rows(
                    min_col=1, max_col=_POINT_COLUMNS, values_only=True
                )
            ]
            try:
                series.append(
                    _read_sheet(
                        rows=rows,
                        name=name,
                        file=path.name,
                        imported_at=imported_at,
                        kind=kind,
                    )
                )
            except HydrofitError as exc:
                problems.append(SheetProblem(name, str(exc)))
    finally:
        workbook.close()

    return ImportResult(tuple(series), tuple(problems))
